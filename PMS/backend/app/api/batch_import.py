"""
批量导入任务 API

支持 Excel/CSV 文件批量创建任务
"""
import io
import csv
import uuid
from datetime import datetime
from typing import Annotated, List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core import get_db
from app.models import User, Task, TaskStatus, TaskType, TaskCategory, AuditModule, AuditAction
from app.api.auth import get_current_user, get_current_manager_or_admin
from app.services import log_audit

router = APIRouter()


# 导入模板列定义
TEMPLATE_COLUMNS = [
    "任务标题",  # title (必填)
    "任务描述",  # description
    "任务类型",  # task_type: performance/daily
    "计划分类",  # category
    "计划开始时间",  # plan_start: YYYY-MM-DD
    "计划完成时间",  # plan_end: YYYY-MM-DD
    "负责人用户名",  # owner_username
    "实施人用户名",  # executor_username
    "权重",  # weight
    "标签",  # tags
]


@router.get("/template")
async def download_template():
    """
    下载导入模板（CSV格式）
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 写入表头
    writer.writerow(TEMPLATE_COLUMNS)
    
    # 写入示例数据
    writer.writerow([
        "示例任务1",
        "这是任务描述",
        "绩效任务",
        "常规类",
        "2026-02-15",
        "2026-02-28",
        "staff",
        "staff",
        "1.0",
        "重要,紧急",
    ])
    writer.writerow([
        "示例任务2",
        "另一个任务",
        "绩效任务",
        "项目类",
        "2026-03-01",
        "2026-03-31",
        "",
        "",
        "1.5",
        "",
    ])
    
    output.seek(0)
    
    # 添加 BOM 以支持 Excel 正确识别 UTF-8
    content = "\ufeff" + output.getvalue()
    
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=task_import_template.csv"}
    )


@router.post("/import")
async def import_tasks(
    file: Annotated[UploadFile, File(description="CSV 或 Excel 文件")],
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_manager_or_admin)],
):
    """
    批量导入任务
    
    支持 CSV 和 Excel (.xlsx) 格式
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="请上传文件")
    
    filename = file.filename.lower()
    
    # 读取文件内容
    content = await file.read()
    
    try:
        if filename.endswith(".csv"):
            rows = parse_csv(content)
        elif filename.endswith(".xlsx"):
            rows = parse_excel(content)
        else:
            raise HTTPException(status_code=400, detail="仅支持 CSV 或 Excel (.xlsx) 格式")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")
    
    if not rows:
        raise HTTPException(status_code=400, detail="文件为空")
    
    # 缓存用户名到ID的映射
    user_map = await get_user_map(db)
    
    # 处理每一行
    results = []
    success_count = 0
    error_count = 0
    
    for idx, row in enumerate(rows, start=2):  # 从第2行开始（第1行是表头）
        async with db.begin_nested():
            try:
                task = await create_task_from_row(db, row, user_map, current_user)
                results.append({
                    "row": idx,
                    "status": "success",
                    "title": row.get("任务标题", ""),
                    "task_id": str(task.id),
                })
                success_count += 1
            except Exception as e:
                # 嵌套事务会自动回滚
                results.append({
                    "row": idx,
                    "status": "error",
                    "title": row.get("任务标题", ""),
                    "message": str(e),
                })
                error_count += 1
    
    # 记录审计日志
    await log_audit(
        db=db,
        user=current_user,
        module=AuditModule.TASK,
        action=AuditAction.TASK_CREATE,
        description=f"批量导入任务: 成功 {success_count} 条, 失败 {error_count} 条",
        details={"success": success_count, "error": error_count, "file": file.filename},
    )
    
    await db.commit()
    
    return {
        "success_count": success_count,
        "error_count": error_count,
        "total": len(rows),
        "details": results,
    }


def parse_csv(content: bytes) -> List[dict]:
    """解析 CSV 文件"""
    # 尝试不同编码
    for encoding in ["utf-8-sig", "utf-8", "gbk", "gb2312"]:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("无法识别文件编码")
    
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def parse_excel(content: bytes) -> List[dict]:
    """解析 Excel 文件"""
    from openpyxl import load_workbook
    
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    
    # 第一行作为表头
    headers = [str(h).strip() if h else "" for h in rows[0]]
    
    result = []
    for row in rows[1:]:
        if not any(row):  # 跳过空行
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = str(val).strip() if val is not None else ""
        result.append(row_dict)
    
    return result


async def get_user_map(db: AsyncSession) -> dict:
    """获取用户名到用户对象的映射"""
    result = await db.execute(select(User).where(User.is_active == True))
    users = result.unique().scalars().all()
    return {u.username: u for u in users}


async def create_task_from_row(
    db: AsyncSession,
    row: dict,
    user_map: dict,
    current_user: User,
) -> Task:
    """从一行数据创建任务"""
    title = row.get("任务标题", "").strip()
    if not title:
        raise ValueError("任务标题不能为空")
    
    # 解析任务类型
    task_type_str = row.get("任务类型", "performance").lower()
    task_type_map = {
        "performance": TaskType.PERFORMANCE,
        "绩效任务": TaskType.PERFORMANCE,
        "project": TaskType.PERFORMANCE,  # 项目任务通常计入绩效
        "专项任务": TaskType.PERFORMANCE,
        "daily": TaskType.DAILY,
        "日常任务": TaskType.DAILY,
    }
    task_type = task_type_map.get(task_type_str, TaskType.PERFORMANCE)
    
    # 解析分类
    category_str = row.get("计划分类", "").strip()
    category_map = {
        "project": TaskCategory.PROJECT,
        "项目类": TaskCategory.PROJECT,
        "routine": TaskCategory.ROUTINE,
        "常规类": TaskCategory.ROUTINE,
        "urgent": TaskCategory.URGENT,
        "紧急类": TaskCategory.URGENT,
        "staged": TaskCategory.STAGED,
        "阶段性": TaskCategory.STAGED,
        "other": TaskCategory.OTHER,
        "其他": TaskCategory.OTHER,
        "日常运维": TaskCategory.ROUTINE,
        "专项工程": TaskCategory.PROJECT,
    }
    category = category_map.get(category_str.lower(), TaskCategory.OTHER)
    
    # 解析日期
    plan_start = parse_date(row.get("计划开始时间", ""))
    plan_end = parse_date(row.get("计划完成时间", ""))
    
    # 解析用户
    owner_username = row.get("负责人用户名", "").strip()
    executor_username = row.get("实施人用户名", "").strip()
    
    owner = user_map.get(owner_username)
    executor = user_map.get(executor_username)
    
    # 解析权重
    weight_str = row.get("权重", "1.0")
    try:
        weight = float(weight_str) if weight_str else 1.0
    except ValueError:
        weight = 1.0
    
    # 创建任务
    task = Task(
        title=title,
        description=row.get("任务描述", ""),
        task_type=task_type,
        category=category,
        tags=row.get("标签", ""),
        plan_start=plan_start,
        plan_end=plan_end,
        owner_id=owner.id if owner else None,
        executor_id=executor.id if executor else None,
        creator_id=current_user.id,
        weight=weight,
        status=TaskStatus.DRAFT,
    )
    
    db.add(task)
    await db.flush()
    
    return task


def parse_date(date_str: str) -> Optional[datetime]:
    """解析日期字符串"""
    if not date_str or not date_str.strip():
        return None
    
    date_str = date_str.strip()
    
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"]:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    
    return None
