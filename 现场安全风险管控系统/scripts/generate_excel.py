import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def generate_risk_system():
    base_path = "/Users/yangjie/Desktop/code/antigravity/现场安全风险管控系统"
    ref_path = os.path.join(base_path, "参考资料")
    output_file = os.path.join(base_path, "作业风险管控系统.xlsx")

    # 1. 准备参数数据 (根据《作业风险评估规范.md》)
    data_s = {
        "序号": ["a", "b", "c", "d", "e", "f"],
        "安全后果": [
            "1人以上死亡或重伤；损大于100万元；一般以上电力事故",
            "2人以上轻伤；损失30万-100万元；三级以上电力事件",
            "1人轻伤；损失20万-30万元；四级电力事件",
            "1人以上轻微伤；损失10万-20万元；五级电力事件",
            "损失1万-10万元；六级电力事件",
            "损失1000元-1万元；七、八级电力事件"
        ],
        "分值": [100, 50, 25, 15, 5, 1]
    }
    
    data_e = {
        "序号": ["a", "b", "c", "d", "e", "f"],
        "频率": ["持续（每天许多次）", "经常（大概每天一次）", "有时（每周一次到每月一次）", "偶尔（每月一次到每年一次）", "很少（据说曾经发生过）", "特别少（没有发生过，但有发生的可能性）"],
        "分值": [10, 6, 3, 2, 1, 0.5]
    }

    data_p = {
        "序号": ["a", "b", "c", "d", "e", "f"],
        "可能性": ["100% 产生预期结果", "十分可能（50%）", "可能（25%）", "很少的可能性", "相当少但确有可能", "百万分之一的可能性"],
        "分值": [10, 6, 3, 1, 0.5, 0.1]
    }

    # 2. 读取原始数据
    try:
        df_risk_list = pd.read_csv(os.path.join(ref_path, "调管电厂涉网二次系统典型作业类型风险清单.csv"), encoding='utf-8')
    except:
        df_risk_list = pd.read_csv(os.path.join(ref_path, "调管电厂涉网二次系统典型作业类型风险清单.csv"), encoding='gbk')

    # 3. 创建 Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 写入参数表
        pd.DataFrame(data_s).to_excel(writer, sheet_name='参数-后果S', index=False)
        pd.DataFrame(data_e).to_excel(writer, sheet_name='参数-暴露E', index=False)
        pd.DataFrame(data_p).to_excel(writer, sheet_name='参数-可能性P', index=False)
        
        # 写入标准风险库
        df_risk_list.to_excel(writer, sheet_name='标准风险库', index=False)
        
        # 创建作业风险管控表 (主界面)
        main_headers = [
            "作业任务", "作业步骤", "危害名称", "危害分布", "特性及产生风险条件", 
            "后果(S)", "暴露(E)", "可能性(P)", "风险值(R)", "风险等级", "控制措施"
        ]
        df_main = pd.DataFrame(columns=main_headers)
        df_main.to_excel(writer, sheet_name='风险管控工具', index=False)

    # 4. 美化与公式 (使用 openpyxl)
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    
    # 格式化函数
    def format_sheet(ws):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 边框
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    for sheet in wb.sheetnames:
        format_sheet(wb[sheet])

    # 在 '风险管控工具' 添加公式和下拉菜单
    ws_main = wb['风险管控工具']
    
    # 设置列宽
    col_widths = {'A': 25, 'B': 20, 'C': 20, 'D': 20, 'E': 30, 'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 15, 'K': 40}
    for col, width in col_widths.items():
        ws_main.column_dimensions[col].width = width

    # 为 F, G, H 列（S, E, P）添加下拉菜单或简单验证（示例：先添加公式）
    # R = S * E * P -> I列 = F * G * H
    for row_num in range(2, 21):  # 预设20行
        # I 列公式
        ws_main[f'I{row_num}'] = f'=F{row_num}*G{row_num}*H{row_num}'
        # J 列等级公式 (嵌套IF)
        ws_main[f'J{row_num}'] = (
            f'=IF(I{row_num}>=400,"特高风险",'
            f'IF(I{row_num}>=200,"高风险",'
            f'IF(I{row_num}>=70,"中等风险",'
            f'IF(I{row_num}>=20,"低风险","可接受风险"))))'
        )
        
        # 条件格式颜色标识 (J列)
        # 注意：openpyxl 的条件格式稍微复杂，这里先手动设置一些基础模拟
        
    # 保存结果
    wb.save(output_file)
    print(f"Excel文件已生成: {output_file}")

if __name__ == "__main__":
    generate_risk_system()
